"""
Export Service

Service for generating Excel and CSV exports with streaming support.
Handles large datasets efficiently.
"""

import io
import csv
from datetime import datetime
from typing import Dict, List, Optional, Any, Union
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.services.trd_buy_service import TrdBuyService
from app.services.lot_service import LotService
from app.services.contract_service import ContractService
from app.services.participant_service import ParticipantService
from app.services.analytics_service import AnalyticsService
import structlog

logger = structlog.get_logger()


class ExportService:
    """
    Export service for generating reports in various formats.
    
    Features:
    - Excel exports with formatting
    - CSV exports for large datasets
    - Streaming for memory efficiency
    - Multiple data sources
    - Custom report templates
    """
    
    def __init__(
        self, 
        session: AsyncSession = None,
        max_rows: int = None,
        chunk_size: int = None,
    ):
        """Initialize Export service."""
        self.session = session
        self.max_rows = max_rows or getattr(settings, 'MAX_EXPORT_ROWS', 100000)
        self.chunk_size = chunk_size or getattr(settings, 'EXPORT_CHUNK_SIZE', 5000)
    
    # Excel Export Methods
    
    async def export_procurement_data(
        self,
        filters: Dict[str, Any] = None,
        format_type: str = "excel",
        include_lots: bool = False,
        include_contracts: bool = False,
        max_rows: int = None,
    ) -> bytes:
        """
        Export procurement data to Excel or CSV.
        
        Args:
            filters: Filter criteria
            format_type: Export format (excel, csv)
            include_lots: Include lot data
            include_contracts: Include contract data
            max_rows: Override default max rows limit
            
        Returns:
            Exported data as bytes
        """
        effective_max_rows = max_rows or self.max_rows
        
        trd_buy_service = TrdBuyService(self.session)
        
        # Get procurement data
        procurement_data = await trd_buy_service.prepare_export_data(
            filters=filters,
            include_lots=include_lots,
            format_for_excel=True,
        )
        
        # Apply row limit
        if len(procurement_data) > effective_max_rows:
            logger.warning(
                "Export data truncated",
                original_rows=len(procurement_data),
                limit=effective_max_rows,
            )
            procurement_data = procurement_data[:effective_max_rows]
        
        if format_type == "excel":
            return await self._create_excel_export(
                data=procurement_data,
                sheet_name="Procurements",
                title="Procurement Data Export",
            )
        else:
            return await self._create_csv_export(procurement_data)
    
    async def export_contract_data(
        self,
        filters: Dict[str, Any] = None,
        format_type: str = "excel",
        include_lot: bool = True,
        include_procurement: bool = False,
        max_rows: int = None,
    ) -> bytes:
        """
        Export contract data to Excel or CSV.
        
        Args:
            filters: Filter criteria
            format_type: Export format
            include_lot: Include lot information
            include_procurement: Include procurement information
            max_rows: Override default max rows limit
            
        Returns:
            Exported data as bytes
        """
        effective_max_rows = max_rows or self.max_rows
        
        contract_service = ContractService(self.session)
        
        # Get contract data
        contract_data = await contract_service.prepare_export_data(
            filters=filters,
            include_lot=include_lot,
            include_procurement=include_procurement,
            format_for_excel=True,
        )
        
        # Apply row limit
        if len(contract_data) > effective_max_rows:
            logger.warning(
                "Export data truncated",
                original_rows=len(contract_data),
                limit=effective_max_rows,
            )
            contract_data = contract_data[:effective_max_rows]
        
        if format_type == "excel":
            return await self._create_excel_export(
                data=contract_data,
                sheet_name="Contracts",
                title="Contract Data Export",
            )
        else:
            return await self._create_csv_export(contract_data)
    
    async def export_participant_data(
        self,
        filters: Dict[str, Any] = None,
        format_type: str = "excel",
        max_rows: int = None,
    ) -> bytes:
        """
        Export participant data to Excel or CSV.
        
        Args:
            filters: Filter criteria
            format_type: Export format
            max_rows: Override default max rows limit
            
        Returns:
            Exported data as bytes
        """
        effective_max_rows = max_rows or self.max_rows
        
        participant_service = ParticipantService(self.session)
        
        # Get participant data
        participant_data = await participant_service.prepare_export_data(
            filters=filters,
            format_for_excel=True,
        )
        
        # Apply row limit
        if len(participant_data) > effective_max_rows:
            logger.warning(
                "Export data truncated",
                original_rows=len(participant_data),
                limit=effective_max_rows,
            )
            participant_data = participant_data[:effective_max_rows]
        
        if format_type == "excel":
            return await self._create_excel_export(
                data=participant_data,
                sheet_name="Participants",
                title="Participant Data Export",
            )
        else:
            return await self._create_csv_export(participant_data)
    
    async def export_analytics_report(
        self,
        report_type: str,
        parameters: Dict[str, Any] = None,
        format_type: str = "excel",
    ) -> bytes:
        """
        Export analytics report.
        
        Args:
            report_type: Type of analytics report
            parameters: Report parameters
            format_type: Export format
            
        Returns:
            Exported data as bytes
        """
        analytics_service = AnalyticsService(self.session)
        
        # Get analytics data
        analytics_data = await analytics_service.prepare_analytics_export(
            report_type=report_type,
            parameters=parameters or {},
        )
        
        if format_type == "excel":
            return await self._create_excel_export(
                data=analytics_data,
                sheet_name="Analytics",
                title=f"Analytics Report - {report_type.replace('_', ' ').title()}",
            )
        else:
            return await self._create_csv_export(analytics_data)
    
    async def export_comprehensive_report(
        self,
        filters: Dict[str, Any] = None,
        include_analytics: bool = True,
        max_rows_per_sheet: int = None,
    ) -> bytes:
        """
        Export comprehensive multi-sheet Excel report.
        
        Args:
            filters: Filter criteria
            include_analytics: Include analytics sheet
            max_rows_per_sheet: Max rows per sheet
            
        Returns:
            Multi-sheet Excel file as bytes
        """
        effective_max_rows = max_rows_per_sheet or self.max_rows
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Export procurement data
        trd_buy_service = TrdBuyService(self.session)
        procurement_data = await trd_buy_service.prepare_export_data(
            filters=filters,
            format_for_excel=True,
        )
        
        if procurement_data:
            if len(procurement_data) > effective_max_rows:
                procurement_data = procurement_data[:effective_max_rows]
            
            await self._add_sheet_to_workbook(
                workbook=workbook,
                data=procurement_data,
                sheet_name="Procurements",
                title="Procurement Data",
            )
        
        # Export contract data
        contract_service = ContractService(self.session)
        contract_data = await contract_service.prepare_export_data(
            filters=filters,
            format_for_excel=True,
        )
        
        if contract_data:
            if len(contract_data) > effective_max_rows:
                contract_data = contract_data[:effective_max_rows]
            
            await self._add_sheet_to_workbook(
                workbook=workbook,
                data=contract_data,
                sheet_name="Contracts",
                title="Contract Data",
            )
        
        # Export participant data
        participant_service = ParticipantService(self.session)
        participant_data = await participant_service.prepare_export_data(
            filters=filters,
            format_for_excel=True,
        )
        
        if participant_data:
            if len(participant_data) > effective_max_rows:
                participant_data = participant_data[:effective_max_rows]
            
            await self._add_sheet_to_workbook(
                workbook=workbook,
                data=participant_data,
                sheet_name="Participants",
                title="Participant Data",
            )
        
        # Add analytics if requested
        if include_analytics:
            analytics_service = AnalyticsService(self.session)
            dashboard_data = await analytics_service.prepare_analytics_export(
                report_type="dashboard_summary",
                parameters=filters or {},
            )
            
            if dashboard_data:
                await self._add_sheet_to_workbook(
                    workbook=workbook,
                    data=dashboard_data,
                    sheet_name="Analytics",
                    title="Analytics Summary",
                )
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info("Comprehensive report generated", sheets=len(workbook.worksheets))
        return output.getvalue()
    
    # Private Helper Methods
    
    async def _create_excel_export(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = "Data",
        title: str = "Export",
    ) -> bytes:
        """Create Excel export from data."""
        if not data:
            # Create empty workbook with message
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            sheet["A1"] = "No data available"
            
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        
        # Create workbook
        workbook = Workbook()
        await self._add_sheet_to_workbook(
            workbook=workbook,
            data=data,
            sheet_name=sheet_name,
            title=title,
        )
        
        # Remove default sheet if it exists
        if "Sheet" in [ws.title for ws in workbook.worksheets]:
            workbook.remove(workbook["Sheet"])
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info("Excel export created", rows=len(data), size_bytes=len(output.getvalue()))
        return output.getvalue()
    
    async def _add_sheet_to_workbook(
        self,
        workbook: Workbook,
        data: List[Dict[str, Any]],
        sheet_name: str,
        title: str,
    ):
        """Add formatted sheet to workbook."""
        if not data:
            return
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create worksheet
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Add title
        worksheet["A1"] = title
        worksheet["A1"].font = Font(size=14, bold=True)
        worksheet["A1"].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet["A1"].font = Font(color="FFFFFF", size=14, bold=True)
        
        # Add metadata
        worksheet["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        worksheet["A3"] = f"Total Records: {len(data)}"
        
        # Add data starting from row 5
        start_row = 5
        
        # Add headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.border = Border(
                top=Side(style="thin"),
                bottom=Side(style="thin"),
                left=Side(style="thin"),
                right=Side(style="thin"),
            )
        
        # Add data rows
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                )
                
                # Format numbers
                if isinstance(value, (int, float, Decimal)):
                    cell.number_format = "#,##0.00" if isinstance(value, (float, Decimal)) else "#,##0"
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _safe_format_value(self, value: Any) -> str:
        """Safely format value for CSV export."""
        if value is None:
            return ""
        elif isinstance(value, datetime):
            return value.isoformat()
        elif isinstance(value, Decimal):
            try:
                return str(float(value))
            except (ValueError, OverflowError):
                return str(value)
        elif isinstance(value, float):
            try:
                return f"{value:.2f}"
            except (ValueError, OverflowError):
                return str(value)
        else:
            return str(value)
    
    async def _create_csv_export(self, data: List[Dict[str, Any]]) -> bytes:
        """Create CSV export from data."""
        if not data:
            return b"No data available\n"
        
        output = io.StringIO()
        
        # Write CSV data
        fieldnames = data[0].keys()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in data:
            # Convert values to strings for CSV with safe formatting
            csv_row = {}
            for key, value in row.items():
                csv_row[key] = self._safe_format_value(value)
            
            writer.writerow(csv_row)
        
        # Convert to bytes
        csv_content = output.getvalue()
        output.close()
        
        logger.info("CSV export created", rows=len(data), size_bytes=len(csv_content.encode()))
        return csv_content.encode("utf-8-sig")  # UTF-8 with BOM for Excel compatibility
    
    # Utility Methods
    
    def get_export_filename(
        self,
        export_type: str,
        format_type: str = "excel",
        timestamp: bool = True,
    ) -> str:
        """
        Generate appropriate filename for export.
        
        Args:
            export_type: Type of export (procurement, contract, etc.)
            format_type: File format
            timestamp: Include timestamp in filename
            
        Returns:
            Generated filename
        """
        base_name = f"scanzakup_{export_type}"
        
        if timestamp:
            timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name += f"_{timestamp_str}"
        
        extension = "xlsx" if format_type == "excel" else "csv"
        return f"{base_name}.{extension}"
    
    def get_export_content_type(self, format_type: str) -> str:
        """Get MIME content type for export format."""
        if format_type == "excel":
            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            return "text/csv"
    
    async def validate_export_request(
        self,
        filters: Dict[str, Any] = None,
        format_type: str = "excel",
        max_rows: int = None,
    ) -> Dict[str, Any]:
        """
        Validate export request and estimate size.
        
        Args:
            filters: Filter criteria
            format_type: Export format
            max_rows: Maximum rows limit
            
        Returns:
            Validation result with estimates
        """
        effective_max_rows = max_rows or self.max_rows
        
        # Estimate row count (this would need to be implemented properly)
        estimated_rows = 1000  # Placeholder
        
        validation = {
            "valid": True,
            "warnings": [],
            "errors": [],
            "estimates": {
                "rows": estimated_rows,
                "size_mb": estimated_rows * 0.001,  # Rough estimate
                "processing_time_seconds": estimated_rows * 0.01,
            },
            "limits": {
                "max_rows": effective_max_rows,
                "chunk_size": self.chunk_size,
            },
        }
        
        # Check limits
        if estimated_rows > effective_max_rows:
            validation["warnings"].append(
                f"Export will be limited to {effective_max_rows} rows (estimated {estimated_rows} total)"
            )
        
        # Check format
        if format_type not in ["excel", "csv"]:
            validation["valid"] = False
            validation["errors"].append("Invalid format type. Must be 'excel' or 'csv'")
        
        return validation 